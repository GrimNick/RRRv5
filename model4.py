import torch
from ultralytics import YOLO
import cv2
import numpy as np
import time
from openpyxl import Workbook
import sys
import os
import subprocess
import pandas as pd  # Import pandas for interpolation

# Load the pre-trained YOLOv8 model
model = YOLO('yolo11n.pt')

# Define classes of interest
VEHICLE_CLASSES = {'motorbike', 'truck', 'bus', 'car' ,'motorcycle'}

def get_class_ids(model, class_names):
    """Retrieve class IDs for the given class names."""
    class_ids = set()
    for key, value in model.names.items():
        if value in class_names:
            class_ids.add(key)
    return class_ids

vehicle_class_ids = get_class_ids(model, VEHICLE_CLASSES)

def get_min_box_area_for_class(class_id):
    """Return the minimum bounding box area based on vehicle type."""
    class_name = model.names[class_id]
    if class_name in {'truck', 'bus'}:
        return 20000
    elif class_name in {'car'}:
        return 10000
    elif class_name in {'motorbike', 'motorcycle'}:
        return 20
    else:
        return 20  # default fallback

# Standalone function to compute Intersection-over-Union (IoU)
def compute_iou(box1, box2):
    x1, y1, x2, y2 = box1
    x3, y3, x4, y4 = box2
    xi1, yi1 = max(x1, x3), max(y1, y3)
    xi2, yi2 = min(x2, x4), min(y2, y4)
    inter_area = max(0, xi2 - xi1) * max(0, yi2 - yi1)
    box1_area = (x2 - x1) * (y2 - y1)
    box2_area = (x4 - x3) * (y4 - y3)
    union_area = box1_area + box2_area - inter_area
    return inter_area / union_area if union_area > 0 else 0

# Merge detections based on center proximity and merge their class info.
def merge_detections(centers, boxes, classes, center_distance_thresh=40):
    merged_centers = []
    merged_boxes = []
    merged_classes = []
    used = [False] * len(centers)
    for i in range(len(centers)):
        if used[i]:
            continue
        current_center = np.array(centers[i], dtype=np.float32)
        current_box = list(boxes[i])
        current_class = classes[i]
        # Use the area of the current box to decide which class to keep
        max_area = (boxes[i][2] - boxes[i][0]) * (boxes[i][3] - boxes[i][1])
        count = 1
        for j in range(i + 1, len(centers)):
            if used[j]:
                continue
            distance = np.linalg.norm(np.array(centers[j]) - np.array(centers[i]))
            if distance < center_distance_thresh:
                current_center = (current_center * count + np.array(centers[j])) / (count + 1)
                current_box[0] = min(current_box[0], boxes[j][0])
                current_box[1] = min(current_box[1], boxes[j][1])
                current_box[2] = max(current_box[2], boxes[j][2])
                current_box[3] = max(current_box[3], boxes[j][3])
                # Update class if this detection has a larger area
                area_j = (boxes[j][2] - boxes[j][0]) * (boxes[j][3] - boxes[j][1])
                if area_j > max_area:
                    max_area = area_j
                    current_class = classes[j]
                count += 1
                used[j] = True
        merged_centers.append(tuple(map(int, current_center)))
        merged_boxes.append(tuple(current_box))
        merged_classes.append(current_class)
    return merged_centers, merged_boxes, merged_classes





class KalmanFilter:
    def __init__(self, x, y):
        self.state = np.array([x, y, 0, 0], dtype='float64')  # Initial position and velocity
        self.F = np.array([[1, 0, 1, 0], 
                           [0, 1, 0, 1], 
                           [0, 0, 1, 0], 
                           [0, 0, 0, 1]])
        self.Q = np.diag([2.0, 2.0, 0.5, 0.5])  # Process noise
        self.H = np.array([[1, 0, 0, 0], 
                           [0, 1, 0, 0]])
        self.R = np.diag([0.1, 0.1])  # Measurement noise
        self.P = np.eye(4) * 200  # Initial uncertainty

    def predict(self):
        self.state = np.dot(self.F, self.state)
        self.P = np.dot(np.dot(self.F, self.P), self.F.T) + self.Q
        return self.state[:2]

    def update(self, measurement):
        z = np.array(measurement)
        y = z - np.dot(self.H, self.state)
        S = np.dot(np.dot(self.H, self.P), self.H.T) + self.R
        K = np.dot(np.dot(self.P, self.H.T), np.linalg.inv(S))
        self.state += np.dot(K, y)
        I = np.eye(self.P.shape[0])
        self.P = np.dot(I - np.dot(K, self.H), self.P)

    def get_velocity(self):
        return np.linalg.norm(self.state[2:])  # Velocity magnitude

class KalmanVehicleTracker:
    def __init__(self, memory_limit=50):  # Increased memory limit for occlusion handling
        self.tracks = {}
        self.memory_limit = memory_limit
        self.next_track_id = 1
        self.lost_tracks = {}
        self.velocity_history = {}

    def iou(self, box1, box2):
        return compute_iou(box1, box2)

    def predict_and_update(self, detections, boxes, classes):
        updated_tracks = {}
        unmatched_detections = list(zip(detections, boxes, classes))

        # First try to match existing tracks
        for track_id, kf in list(self.tracks.items()):
            predicted_pos = kf.predict()
            best_match = None
            best_iou = 0.0

            for detection, box, cls in unmatched_detections:
                pred_box = [predicted_pos[0] - 10, predicted_pos[1] - 10, 
                            predicted_pos[0] + 10, predicted_pos[1] + 10]
                iou_score = self.iou(box, pred_box)
                distance = np.linalg.norm(np.array(predicted_pos) - np.array(detection))
                if iou_score > best_iou and distance < 50:
                    best_iou = iou_score
                    best_match = (detection, box, cls)
            if best_match:
                kf.update(best_match[0])
                updated_tracks[track_id] = kf
                unmatched_detections.remove(best_match)
            else:
                self.lost_tracks[track_id] = (kf, self.memory_limit)

        # Attempt re-association for lost tracks
        for track_id, (kf, frames_left) in list(self.lost_tracks.items()):
            predicted_pos = kf.predict()
            best_match = None
            best_iou = 0.0
            for detection, box, cls in unmatched_detections:
                pred_box = [predicted_pos[0] - 10, predicted_pos[1] - 10,
                            predicted_pos[0] + 10, predicted_pos[1] + 10]
                iou_score = self.iou(box, pred_box)
                distance = np.linalg.norm(np.array(predicted_pos) - np.array(detection))
                if iou_score > best_iou and distance < 50:
                    best_iou = iou_score
                    best_match = (detection, box, cls)
            if best_match:
                kf.update(best_match[0])
                updated_tracks[track_id] = kf
                unmatched_detections.remove(best_match)
                del self.lost_tracks[track_id]
            else:
                if frames_left > 0:
                    self.lost_tracks[track_id] = (kf, frames_left - 1)
                else:
                    del self.lost_tracks[track_id]

        # Create new tracks for remaining unmatched detections if box area is above threshold.
        for detection, box, cls in unmatched_detections:
            box_area = (box[2] - box[0]) * (box[3] - box[1])
            min_area = get_min_box_area_for_class(cls)
            if box_area < min_area:
                # Likely a part of a vehicle; do not register as a new track.
                continue
            kf = KalmanFilter(detection[0], detection[1])
            updated_tracks[self.next_track_id] = kf
            self.next_track_id += 1

        self.tracks = updated_tracks
        return self.tracks

tracker = KalmanVehicleTracker(memory_limit=50)

def format_timestamp(frame_count, fps):
    total_seconds = frame_count / fps
    hours = int(total_seconds // 3600)
    minutes = int((total_seconds % 3600) // 60)
    seconds = int(total_seconds % 60)
    return f"{hours:02}:{minutes:02}:{seconds:02}"
    

def process_video(video_path):
    cap = cv2.VideoCapture(video_path)
    fps = cap.get(cv2.CAP_PROP_FPS)
    width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
    height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))

    fourcc = cv2.VideoWriter_fourcc(*'avc1')
    output_path = video_path.replace('.mp4', '_processed.mp4')
    out = cv2.VideoWriter(output_path, fourcc, fps, (width, height))
    output_file_name = output_path.replace('.mp4', '_data.xlsx')
    wb = Workbook()
    frame_count = 0
    camera_height = 13
    camera_angle = -52
    total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
    print(f"Total Frames in Input:{total_frames}")
    if not hasattr(process_video, "vehicle_first_seen"):
        process_video.vehicle_first_seen = {}

    while cap.isOpened():
        ret, frame = cap.read()
        if not ret:
            break

        results = model(frame ,conf= 0.25,imgsz= 1280) 
        detected_objects = []
        detected_boxes = []
        detected_classes = []


        for box in results[0].boxes:
            x1, y1, x2, y2 = map(int, box.xyxy[0])
            class_id = int(box.cls[0])

            if class_id in vehicle_class_ids:
                center_x = (x1 + x2) // 2
                center_y = (y1 + y2) // 2
                detected_objects.append((center_x, center_y))
                detected_boxes.append((x1, y1, x2, y2))
                detected_classes.append(class_id)
                # Merge overlapping detections (if any) delete if necessary
        if detected_objects and detected_boxes and detected_classes:
            detected_objects, detected_boxes, detected_classes = merge_detections(
                detected_objects, detected_boxes, detected_classes, center_distance_thresh=40
            )


        tracks = tracker.predict_and_update(detected_objects, detected_boxes,detected_classes)

        timestamp = format_timestamp(frame_count, fps)

        velocity_data = []  # Store the velocity for interpolation
        current_time_sec = frame_count / fps

        for track_id, kf in tracks.items():
            pred_x, pred_y = kf.state[:2]
            velocity = kf.get_velocity()
            scaled_velocity = velocity / (np.cos(np.radians(camera_angle)) * camera_height)
            velocity = max(scaled_velocity - 0.5, 0)

            velocity_data.append((timestamp, track_id, velocity))

            sheet_name = f'Track ID {track_id}'
            if sheet_name not in wb.sheetnames:
                ws = wb.create_sheet(sheet_name)
                ws.append(["Time", "Velocity", "Relative Velocity"])
            else:
                ws = wb[sheet_name]

    # Count the number of vehicles in the current frame
        for track_id in tracks.keys():
            if track_id not in process_video.vehicle_first_seen:
                process_video.vehicle_first_seen[track_id] = current_time_sec
        num_vehicles_total = len([tid for tid, first_seen in process_video.vehicle_first_seen.items() 
                                  if (current_time_sec - first_seen) >= 1.0])
        num_vehicles_current_frame = len(tracks)

        # Display text on the video frame
        text1 = f"Number of vehicles in current frame: {num_vehicles_current_frame}"
        text2 = f"Number of vehicles till now: {num_vehicles_total}"
        
        # Put text at bottom right of the video
        text_x = width - 450
        text_y = height - 50
        
        cv2.putText(frame, text1, (text_x, text_y), 
                    cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 255), 2)
        cv2.putText(frame, text2, (text_x, text_y + 25), 
                    cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 255), 2)
        print(f"Frame {frame_count}: Current vehicles={num_vehicles_current_frame}, Total detected={num_vehicles_total}")


        for track_id, kf in tracks.items():
            pred_x, pred_y = kf.state[:2]
            velocity = kf.get_velocity()

            for (center_x, center_y), (x1, y1, x2, y2) in zip(detected_objects, detected_boxes):
                if abs(pred_x - center_x) < 30 and abs(pred_y - center_y) < 30:
                    cv2.rectangle(frame, (x1, y1), (x2, y2), color=(0, 255, 0), thickness=2)
                    cv2.putText(frame, f'ID: {track_id} Vel: {velocity:.2f}', (x1, y1 - 10),
                                cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 255, 0), 2)
                    break

        # Interpolate missing velocity values
        velocity_df = pd.DataFrame(velocity_data, columns=["Time", "Track ID", "Velocity"])
        velocity_df['Velocity'] = velocity_df['Velocity'].replace(0, np.nan)  # Replace 0 with NaN for interpolation
        velocity_df['Velocity'] = velocity_df['Velocity'].interpolate(method='linear')  # Apply interpolation
        
        # Re-append interpolated velocities back to the Excel sheet
        for _, row in velocity_df.iterrows():
            track_id = int(row["Track ID"])
            ts = row["Time"]
            vel = row["Velocity"]
            sheet_name = f'Track ID {track_id}'
            # Create a new sheet if it does not exist
            ws = wb[sheet_name]
            relative_velocity = max(np.mean([np.linalg.norm(other_kf.state[2:]) 
                                             for other_id, other_kf in tracks.items() if other_id != track_id]), 0)
            ws.append([ts, vel, relative_velocity])
        
        out.write(frame)
        frame_count += 1
        perc = (frame_count / total_frames) * 100
        if frame_count % 10 == 0:
            print(f"Processing frame {frame_count}/{total_frames} - {perc:.2f}% completed", flush=True)

    cap.release()
    out.release()
    wb.save(output_file_name)
    wb.close()
    print("Processing complete.", flush=True)
    full_output_file_path = os.path.abspath(output_file_name)
    subprocess.run([sys.executable, 'modelExcel.py', full_output_file_path])
    
if __name__ == "__main__":
    if len(sys.argv) > 1:
        process_video(sys.argv[1])
    else:
        print("No video path provided.")
